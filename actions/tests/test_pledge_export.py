from __future__ import annotations

import csv
import io
import uuid
from io import BytesIO
from typing import TYPE_CHECKING

from django.db.models import Count
from django.http import StreamingHttpResponse
from wagtail.models import Locale

import pytest

from actions.models import Pledge, PledgeCommitment, PublicUser
from actions.pledge_admin import PledgeIndexView, PledgeViewSet
from actions.tests.factories import PlanFactory, PledgeFactory
from actions.tests.utils import instantiate_view

if TYPE_CHECKING:
    from django.test import RequestFactory

    from users.models import User

pytestmark = pytest.mark.django_db


def _get_view_and_queryset(rf: RequestFactory, user: User, plan, extra_params: dict | None = None):
    """Set up a PledgeIndexView and build the queryset for the given plan."""
    view_set = PledgeViewSet()
    view = instantiate_view(view_set.index_view, PledgeIndexView)
    params = {'export': 'csv'}
    if extra_params:
        params.update(extra_params)
    request = rf.get('/admin/', params)
    request.user = user
    view.setup(request)
    view.filterset_class = None  # type: ignore[assignment]
    queryset = Pledge.objects.filter(plan=plan).annotate(commitment_count=Count('commitments'))
    return view, queryset


def _parse_csv_rows(view: PledgeIndexView, queryset) -> list[dict[str, str]]:
    """Stream pledges CSV from the view and parse into a list of row dicts."""
    raw = b''.join(view.stream_csv(queryset))
    reader = csv.DictReader(io.StringIO(raw.decode('utf-8')))
    return list(reader)


def _parse_commitments_csv_rows(view: PledgeIndexView, queryset) -> list[dict[str, str]]:
    """Stream commitments CSV from the view and parse into a list of row dicts."""
    response = view.write_commitments_csv_response(queryset)
    raw = b''.join(response.streaming_content)  # type: ignore[arg-type]
    reader = csv.DictReader(io.StringIO(raw.decode('utf-8')))
    return list(reader)


class TestPledgeExportCSV:
    @pytest.fixture(autouse=True)
    def setup(self, plan_admin_user):
        self.user = plan_admin_user
        self.plan = self.user.get_active_admin_plan()
        self.plan.features.enable_community_engagement = True
        self.plan.features.save()

    def test_basic_export_columns(self, rf):
        """Pledge export should contain base columns; user_data columns should not appear."""
        PledgeFactory.create(plan=self.plan)

        view, qs = _get_view_and_queryset(rf, self.user, self.plan)
        rows = _parse_csv_rows(view, qs)

        assert len(rows) == 1
        row = rows[0]
        assert set(row.keys()) >= {'ID', 'Name', 'Slug', 'Number of commitments'}
        assert not any(k in row for k in ('zip_code', 'city'))

    def test_commitment_count(self, rf):
        """Pledge export should reflect the correct number of commitments per pledge."""
        pledge = PledgeFactory.create(plan=self.plan)
        for _ in range(3):
            PledgeCommitment.objects.create(
                pledge=pledge,
                public_user=PublicUser.objects.create(),
            )

        view, qs = _get_view_and_queryset(rf, self.user, self.plan)
        rows = _parse_csv_rows(view, qs)

        assert rows[0]['Number of commitments'] == '3'

    def test_user_data_columns_absent_from_pledge_export(self, rf):
        """User data keys from commitments should NOT appear as columns in the pledge export."""
        pledge = PledgeFactory.create(plan=self.plan)
        PledgeCommitment.objects.create(
            pledge=pledge,
            public_user=PublicUser.objects.create(user_data={'zip_code': '00100', 'city': 'Helsinki'}),
        )

        view, qs = _get_view_and_queryset(rf, self.user, self.plan)
        rows = _parse_csv_rows(view, qs)

        row = rows[0]
        assert 'zip_code' not in row
        assert 'city' not in row

    def test_multiple_pledges(self, rf):
        """Pledge export should include one row per pledge."""
        PledgeFactory.create(plan=self.plan, name='First Pledge')
        PledgeFactory.create(plan=self.plan, name='Second Pledge')

        view, qs = _get_view_and_queryset(rf, self.user, self.plan)
        rows = _parse_csv_rows(view, qs)

        names = {row['Name'] for row in rows}
        assert names == {'First Pledge', 'Second Pledge'}

    def test_only_own_plan_pledges(self, rf):
        """Pledge export should only include pledges from the user's active plan."""
        PledgeFactory.create(plan=self.plan, name='My Pledge')

        other_plan = PlanFactory.create()
        other_plan.features.enable_community_engagement = True
        other_plan.features.save()
        PledgeFactory.create(plan=other_plan, name='Other Pledge')

        view, qs = _get_view_and_queryset(rf, self.user, self.plan)
        rows = _parse_csv_rows(view, qs)

        names = {row['Name'] for row in rows}
        assert 'My Pledge' in names
        assert 'Other Pledge' not in names


class TestCommitmentsExportCSV:
    @pytest.fixture(autouse=True)
    def setup(self, plan_admin_user):
        self.user = plan_admin_user
        self.plan = self.user.get_active_admin_plan()
        self.plan.features.enable_community_engagement = True
        self.plan.features.save()

    def test_commitments_export_columns(self, rf):
        """Commitments export should contain the expected base column headers."""
        pledge = PledgeFactory.create(plan=self.plan)
        PledgeCommitment.objects.create(pledge=pledge, public_user=PublicUser.objects.create())

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export_type': 'commitments'})
        rows = _parse_commitments_csv_rows(view, qs)

        assert len(rows) == 1
        assert set(rows[0].keys()) >= {'Pledge ID', 'Pledge name', 'Commitment date', 'User ID'}

    def test_commitments_one_row_per_commitment(self, rf):
        """Commitments export should yield one row per commitment, not per pledge."""
        pledge = PledgeFactory.create(plan=self.plan)
        for _ in range(3):
            PledgeCommitment.objects.create(
                pledge=pledge,
                public_user=PublicUser.objects.create(),
            )

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export_type': 'commitments'})
        rows = _parse_commitments_csv_rows(view, qs)

        assert len(rows) == 3

    def test_commitments_user_data_in_own_cells(self, rf):
        """Each commitment's user_data values should appear in their own cells, not comma-joined."""
        pledge = PledgeFactory.create(plan=self.plan)
        PledgeCommitment.objects.create(
            pledge=pledge,
            public_user=PublicUser.objects.create(user_data={'zip_code': '00100', 'city': 'Helsinki'}),
        )
        PledgeCommitment.objects.create(
            pledge=pledge,
            public_user=PublicUser.objects.create(user_data={'zip_code': '00200'}),
        )

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export_type': 'commitments'})
        rows = _parse_commitments_csv_rows(view, qs)

        assert len(rows) == 2
        zip_codes = {row['zip_code'] for row in rows}
        assert zip_codes == {'00100', '00200'}
        # Each cell contains a single value, not a comma-joined list
        for row in rows:
            assert ',' not in row['zip_code']

    def test_commitments_includes_timestamp_and_user_id(self, rf):
        """Each commitment row should include a timestamp and the pledge user's UUID."""
        pledge = PledgeFactory.create(plan=self.plan)
        pu = PublicUser.objects.create()
        PledgeCommitment.objects.create(pledge=pledge, public_user=pu)

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export_type': 'commitments'})
        rows = _parse_commitments_csv_rows(view, qs)

        assert len(rows) == 1
        row = rows[0]
        assert row['User ID'] == str(pu.uuid)
        # Timestamp should be an ISO 8601 string
        assert 'T' in row['Commitment date'] or '-' in row['Commitment date']

    def test_commitments_only_own_plan(self, rf):
        """Commitments export should only include commitments from the user's active plan."""
        pledge = PledgeFactory.create(plan=self.plan, name='My Pledge')
        PledgeCommitment.objects.create(pledge=pledge, public_user=PublicUser.objects.create())

        other_plan = PlanFactory.create()
        other_plan.features.enable_community_engagement = True
        other_plan.features.save()
        other_pledge = PledgeFactory.create(plan=other_plan, name='Other Pledge')
        PledgeCommitment.objects.create(pledge=other_pledge, public_user=PublicUser.objects.create())

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export_type': 'commitments'})
        rows = _parse_commitments_csv_rows(view, qs)

        pledge_names = {row['Pledge name'] for row in rows}
        assert 'My Pledge' in pledge_names
        assert 'Other Pledge' not in pledge_names

    def test_commitments_comma_in_user_data_value(self, rf):
        """User data values containing commas should be properly handled in the commitments CSV."""
        pledge = PledgeFactory.create(plan=self.plan)
        PledgeCommitment.objects.create(
            pledge=pledge,
            public_user=PublicUser.objects.create(user_data={'location': 'City, State'}),
        )

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export_type': 'commitments'})
        rows = _parse_commitments_csv_rows(view, qs)

        assert rows[0]['location'] == 'City, State'


def _parse_xlsx_workbook(view: PledgeIndexView, queryset):
    """Call write_xlsx_response and return an openpyxl Workbook for inspection."""
    import openpyxl  # type: ignore[import-untyped]

    response = view.write_xlsx_response(queryset)
    raw = b''.join(response.streaming_content)  # type: ignore[arg-type]
    return openpyxl.load_workbook(BytesIO(raw))


class TestCombinedXlsxExport:
    @pytest.fixture(autouse=True)
    def setup(self, plan_admin_user):
        self.user = plan_admin_user
        self.plan = self.user.get_active_admin_plan()
        self.plan.features.enable_community_engagement = True
        self.plan.features.save()

    def test_xlsx_sheet_names(self, rf):
        """XLSX export should contain 'Pledges' and 'Commitments' sheets."""
        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export': 'xlsx'})
        wb = _parse_xlsx_workbook(view, qs)

        assert 'Pledges' in wb.sheetnames
        assert 'Commitments' in wb.sheetnames

    def test_xlsx_pledges_sheet_one_row_per_pledge(self, rf):
        """Pledges sheet should have one data row per pledge."""
        PledgeFactory.create(plan=self.plan, name='Alpha')
        PledgeFactory.create(plan=self.plan, name='Beta')

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export': 'xlsx'})
        wb = _parse_xlsx_workbook(view, qs)

        data_rows = list(wb['Pledges'].iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2

    def test_xlsx_commitments_sheet_one_row_per_commitment(self, rf):
        """Commitments sheet should have one data row per commitment."""
        pledge = PledgeFactory.create(plan=self.plan)
        for _ in range(3):
            PledgeCommitment.objects.create(pledge=pledge, public_user=PublicUser.objects.create())

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export': 'xlsx'})
        wb = _parse_xlsx_workbook(view, qs)

        data_rows = list(wb['Commitments'].iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 3

    def test_xlsx_user_data_in_own_columns(self, rf):
        """User data values should appear in separate named columns in the Commitments sheet."""
        pledge = PledgeFactory.create(plan=self.plan)
        PledgeCommitment.objects.create(
            pledge=pledge,
            public_user=PublicUser.objects.create(user_data={'zip_code': '00100', 'city': 'Helsinki'}),
        )

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export': 'xlsx'})
        wb = _parse_xlsx_workbook(view, qs)

        ws = wb['Commitments']
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert 'zip_code' in header
        assert 'city' in header
        zip_idx = header.index('zip_code')
        data_row = next(iter(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
        assert data_row[zip_idx] == '00100'

    def test_render_to_response_routes_commitments_csv(self, rf):
        """render_to_response with export_type=commitments and export=csv returns a streaming CSV."""
        pledge = PledgeFactory.create(plan=self.plan)
        PledgeCommitment.objects.create(pledge=pledge, public_user=PublicUser.objects.create())

        view, qs = _get_view_and_queryset(rf, self.user, self.plan, {'export': 'csv', 'export_type': 'commitments'})
        response = view.render_to_response({'object_list': qs})

        assert isinstance(response, StreamingHttpResponse)
        assert 'text/csv' in response['Content-Type']


class TestPledgeIndexLocaleFiltering:
    @pytest.fixture(autouse=True)
    def setup(self, plan_admin_user):
        self.user = plan_admin_user
        self.plan = self.user.get_active_admin_plan()
        self.plan.primary_language = 'en'
        self.plan.other_languages = ['fi']
        self.plan.features.enable_community_engagement = True
        self.plan.save(update_fields=['primary_language', 'other_languages'])
        self.plan.features.save()

    def test_index_queryset_includes_only_primary_locale(self, rf):
        """Pledge index queryset should include only primary-language pledges."""
        primary_pledge = PledgeFactory.create(
            plan=self.plan,
            name='English pledge',
            slug='pledge-locale-filter',
        )
        fi_locale, _ = Locale.objects.get_or_create(language_code='fi')
        fi_pledge = primary_pledge.copy_for_translation(fi_locale)
        fi_pledge.uuid = uuid.uuid4()
        fi_pledge.name = 'Suomenkielinen lupaus'
        fi_pledge.save()

        request = rf.get('/admin/')
        request.user = self.user
        queryset = PledgeViewSet().get_queryset(request)

        ids = set(queryset.values_list('id', flat=True))
        assert primary_pledge.id in ids
        assert fi_pledge.id not in ids
